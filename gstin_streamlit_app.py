"""
GSTIN Compliance Suite
=======================

A Streamlit dashboard for bulk GSTIN profile lookups, public GST return
filing tracking, and OTP-authorised GSTR-1 / GSTR-3B data retrieval via the
Sandbox (api.sandbox.co.in) GST Compliance APIs.

This tool never collects or stores GST portal passwords. Private return data
is only reachable once the taxpayer (or their authorised user) completes an
OTP consent flow for that specific GSTIN — exactly as required by
GSTN-compliant GST Suvidha Provider integrations.
"""

import html
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import plotly.express as px
import requests
import streamlit as st
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

# ============================================================
# App Config
# ============================================================
PROD_BASE_URL = "https://api.sandbox.co.in"
TEST_BASE_URL = "https://test-api.sandbox.co.in"

AUTH_ENDPOINT = "/authenticate"

# Public API endpoints
SEARCH_GSTIN_ENDPOINT = "/gst/compliance/public/gstin/search"
TRACK_GSTR_ENDPOINT = "/gst/compliance/public/gstrs/track"
PREFERENCE_ENDPOINT = "/gst/compliance/public/gstrs/preference"

# Taxpayer-private API endpoints
TAXPAYER_OTP_ENDPOINT = "/gst/compliance/tax-payer/otp"
TAXPAYER_OTP_VERIFY_ENDPOINT = "/gst/compliance/tax-payer/otp/verify"
GSTR3B_DETAILS_ENDPOINT_TEMPLATE = "/gst/compliance/tax-payer/gstrs/gstr-3b/{year}/{month}"

# GSTR-1 document sections.
# Endpoint pattern: /gst/compliance/tax-payer/gstrs/gstr-1/{section}/{year}/{month}
GSTR1_SECTIONS = {
    "B2B": "b2b",
    "B2BA": "b2ba",
    "B2CL": "b2cl",
    "B2CLA": "b2cla",
    "B2CS": "b2cs",
    "B2CSA": "b2csa",
    "CDNR": "cdnr",
    "CDNRA": "cdnra",
    "CDNUR": "cdnur",
    "CDNURA": "cdnura",
    "EXP": "exp",
    "EXPA": "expa",
    "AT": "at",
    "ATA": "ata",
    "TXP": "txp",
    "TXPA": "txpa",
    "NIL": "nil",
    "HSN": "hsn",
    "DOC-ISSUE": "doc-issue",
}
COMMON_GSTR1_SECTIONS = ["B2B", "B2CL", "B2CS", "CDNR", "CDNUR", "EXP", "NIL", "HSN", "DOC-ISSUE"]

DEFAULT_TIMEOUT = 60
MAX_RETRIES = 2
RETRYABLE_STATUS_CODES = {429, 502, 503, 504}

STEP_VALIDATION = "Validation"
STEP_PROFILE = "GSTIN Search"
STEP_PREFERENCE = "Return Preference"
STEP_FILINGS = "Track GST Returns"

SUCCESS_API_STATUSES = {"OK", "Profile OK", "Completed"}
FAILURE_API_STATUSES = {"Skipped", "Warning", "Profile Failed", "Preference Failed", "Filing Failed", "Failed"}

STATE_CODES = {
    "01": "Jammu & Kashmir",
    "02": "Himachal Pradesh",
    "03": "Punjab",
    "04": "Chandigarh",
    "05": "Uttarakhand",
    "06": "Haryana",
    "07": "Delhi",
    "08": "Rajasthan",
    "09": "Uttar Pradesh",
    "10": "Bihar",
    "11": "Sikkim",
    "12": "Arunachal Pradesh",
    "13": "Nagaland",
    "14": "Manipur",
    "15": "Mizoram",
    "16": "Tripura",
    "17": "Meghalaya",
    "18": "Assam",
    "19": "West Bengal",
    "20": "Jharkhand",
    "21": "Odisha",
    "22": "Chhattisgarh",
    "23": "Madhya Pradesh",
    "24": "Gujarat",
    "25": "Daman & Diu",
    "26": "Dadra & Nagar Haveli",
    "27": "Maharashtra",
    "28": "Andhra Pradesh",
    "29": "Karnataka",
    "30": "Goa",
    "31": "Lakshadweep",
    "32": "Kerala",
    "33": "Tamil Nadu",
    "34": "Puducherry",
    "35": "Andaman & Nicobar Islands",
    "36": "Telangana",
    "37": "Andhra Pradesh",
    "38": "Ladakh",
    "97": "Other Territory",
    "99": "Centre Jurisdiction",
}

COL_GSTIN = "GSTIN"
COL_GSTIN_STATUS = "GSTIN / UIN Status"
COL_API_STATUS = "API Status"

CUSTOMER_COLUMNS = [
    "GSTIN",
    "Valid Format",
    "Checksum Valid",
    "State From GSTIN",
    "PAN From GSTIN",
    "Legal Name of Business",
    "Trade Name",
    "Constitution of Business",
    "GSTIN / UIN Status",
    "Taxpayer Type",
    "Registration Date",
    "Cancellation Date",
    "Last Updated on GSTN",
    "E-Invoice Status",
    "Aggregate Turnover",
    "Aggregate Turnover FY",
    "Gross Total Income",
    "Gross Total Income FY",
    "Turnover Data Status",
    "Detected Turnover/E-Invoice Keys",
    "Raw GSTIN Profile JSON",
    "Nature of Business",
    "Principal Place State",
    "Principal Place City",
    "Principal Place Pincode",
    "Filing Frequency",
    "Preference Detail",
    "API Status",
    "API Message",
    "Fetched At",
]

FILING_COLUMNS = [
    "GSTIN",
    "Return Type",
    "Return Period",
    "Date of Filing",
    "Status",
    "ARN",
    "Mode of Filing",
    "Valid",
    "Financial Year",
]

ERROR_COLUMNS = [
    "GSTIN",
    "Step",
    "Category",
    "Error",
    "Raw Response",
    "Fetched At",
]

RETURN_RESPONSE_COLUMNS = [
    "GSTIN",
    "Return",
    "Section",
    "Year",
    "Month",
    "Status",
    "Message",
    "Transaction ID",
    "Fetched At",
    "Raw JSON",
]

RETURN_FACT_COLUMNS = [
    "GSTIN",
    "Return",
    "Section",
    "Year",
    "Month",
    "Path",
    "Value",
]

GSTR1_INVOICE_COLUMNS = [
    "GSTIN",
    "Year",
    "Month",
    "Section",
    "Counterparty GSTIN",
    "Invoice Number",
    "Invoice Date",
    "Invoice Value",
    "Invoice Type",
    "POS",
    "Reverse Charge",
    "Taxable Value",
    "IGST",
    "CGST",
    "SGST",
    "CESS",
    "Source Type",
    "Raw Path",
]

ACCENTS = {
    "Saffron": {"accent": "#C2730A", "accent_soft": "#FCEEDC"},
    "Indigo": {"accent": "#1E3A8A", "accent_soft": "#E4E9F8"},
    "Teal": {"accent": "#0F766E", "accent_soft": "#DFF3F1"},
    "Rose": {"accent": "#B91C5C", "accent_soft": "#FBE3EE"},
}
CHART_PALETTE = ["#1E3A8A", "#0F766E", "#C2730A", "#7C879F", "#B91C5C", "#334155"]

STATUS_ICON_RULES = [
    ("cancelled", "🔴"),
    ("canceled", "🔴"),
    ("inactive", "🔴"),
    ("failed", "🔴"),
    ("active", "🟢"),
    ("suspended", "🟠"),
    ("warning", "🟡"),
    ("skipped", "⚪"),
    ("not started", "⚪"),
    ("completed", "✅"),
]

RESETTABLE_KEYS = [
    "gstin_list",
    "customers_df",
    "filings_df",
    "errors_df",
    "taxpayer_sessions",
    "pending_taxpayer_auth",
    "return_responses_df",
    "return_facts_df",
    "gstr1_invoices_df",
    "activity_log",
]


# ============================================================
# GSTIN Validation
# ============================================================
def clean_gstin(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z]", "", str(value or "")).upper()


def gstin_format_valid(gstin: str) -> bool:
    gstin = clean_gstin(gstin)
    pattern = r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$"
    return bool(re.match(pattern, gstin))


def compute_gstin_check_digit(first_14_chars: str) -> str:
    chars = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    factor = 2
    total = 0

    for char in reversed(first_14_chars.upper()):
        code_point = chars.find(char)
        if code_point == -1:
            return ""

        addend = factor * code_point
        factor = 1 if factor == 2 else 2
        addend = (addend // 36) + (addend % 36)
        total += addend

    check_code_point = (36 - (total % 36)) % 36
    return chars[check_code_point]


def gstin_checksum_valid(gstin: str) -> bool:
    gstin = clean_gstin(gstin)
    if len(gstin) != 15:
        return False
    return compute_gstin_check_digit(gstin[:14]) == gstin[-1]


def gstin_state(gstin: str) -> str:
    gstin = clean_gstin(gstin)
    return STATE_CODES.get(gstin[:2], "Unknown State Code")


def pan_from_gstin(gstin: str) -> str:
    gstin = clean_gstin(gstin)
    return gstin[2:12] if len(gstin) >= 12 else ""


def gstin_segments(gstin: str) -> Dict[str, str]:
    g = clean_gstin(gstin)
    return {
        "state": g[0:2] if len(g) >= 2 else "",
        "pan": g[2:12] if len(g) >= 12 else g[2:],
        "entity": g[12:13] if len(g) >= 13 else "",
        "z": g[13:14] if len(g) >= 14 else "",
        "checksum": g[14:15] if len(g) >= 15 else "",
    }


def parse_period_to_year_month(return_period: str) -> Tuple[str, str]:
    """
    Converts GST return period like 052026 into (year=2026, month=05).
    """
    value = re.sub(r"[^0-9]", "", str(return_period or ""))
    if len(value) == 6:
        month = value[:2]
        year = value[2:]
        return year, month
    return "", ""


# ============================================================
# Common API Helpers
# ============================================================
def get_base_url(env: str) -> str:
    return PROD_BASE_URL if env.lower() == "production" else TEST_BASE_URL


def get_nested(data: Dict[str, Any], *keys: str, default: Any = None) -> Any:
    current: Any = data
    for key in keys:
        if not isinstance(current, dict):
            return default
        current = current.get(key)
    return current if current is not None else default


def classify_error(message: str) -> str:
    text = (message or "").lower()
    if "401" in text or "unauthor" in text or "access_token" in text:
        return "Authentication"
    if "403" in text or "forbidden" in text:
        return "Permission"
    if "429" in text or "rate limit" in text or "too many requests" in text:
        return "Rate Limited"
    if "timed out" in text or "timeout" in text:
        return "Timeout"
    if "connection" in text or "network" in text:
        return "Network"
    if "checksum" in text or "invalid gstin" in text or "validation" in text:
        return "Validation"
    return "Other"


def _request_json(
    method: str,
    url: str,
    headers: Dict[str, str],
    json_body: Optional[Dict[str, Any]] = None,
    params: Optional[Dict[str, Any]] = None,
    max_retries: int = MAX_RETRIES,
) -> Dict[str, Any]:
    """
    Shared HTTP core for every API call in this app. Retries network errors
    and a small set of transient HTTP status codes (429/502/503/504) with
    capped exponential backoff before giving up.
    """
    last_exc: Optional[Exception] = None

    for attempt in range(max_retries + 1):
        try:
            response = requests.request(
                method,
                url,
                headers=headers,
                json=json_body,
                params=params or {},
                timeout=DEFAULT_TIMEOUT,
            )
        except requests.exceptions.RequestException as exc:
            last_exc = exc
            if attempt < max_retries:
                time.sleep(min(2 ** attempt, 8))
                continue
            raise RuntimeError(f"Network error after {attempt + 1} attempt(s): {exc}") from exc

        try:
            payload = response.json()
        except Exception:
            if response.status_code >= 400:
                raise RuntimeError(f"HTTP {response.status_code}: {response.text[:500]}")
            raise RuntimeError(f"Invalid JSON response: {response.text[:500]}")

        if response.status_code >= 400:
            if response.status_code in RETRYABLE_STATUS_CODES and attempt < max_retries:
                last_exc = RuntimeError(f"HTTP {response.status_code}: {payload}")
                time.sleep(min(2 ** attempt, 8))
                continue
            raise RuntimeError(f"HTTP {response.status_code}: {payload}")

        return payload

    if last_exc:
        raise last_exc
    raise RuntimeError("Request failed for an unknown reason.")


def authenticate(api_key: str, api_secret: str, base_url: str) -> str:
    """
    Authenticates your Sandbox/API-provider account.
    This is NOT GST portal username/password.
    """
    headers = {
        "x-api-key": (api_key or "").strip(),
        "x-api-secret": (api_secret or "").strip(),
    }
    payload = _request_json("POST", f"{base_url}{AUTH_ENDPOINT}", headers, max_retries=1)

    token = get_nested(payload, "data", "access_token")
    if not token:
        raise RuntimeError(f"Authentication succeeded but access_token was not found: {payload}")

    return token


def make_app_headers(api_key: str, token: str, accept_cache: bool = True) -> Dict[str, str]:
    headers = {
        "x-api-key": (api_key or "").strip(),
        "authorization": (token or "").strip(),
        "x-api-version": "1.0",
        "Content-Type": "application/json",
    }
    if accept_cache:
        headers["x-accept-cache"] = "true"
    return headers


def make_taxpayer_headers(api_key: str, taxpayer_token: str) -> Dict[str, str]:
    return {
        "x-api-key": (api_key or "").strip(),
        "authorization": (taxpayer_token or "").strip(),
        "x-api-version": "1.0.0",
        "Content-Type": "application/json",
    }


def post_api(
    base_url: str,
    endpoint: str,
    api_key: str,
    token: str,
    body: Dict[str, Any],
    params: Optional[Dict[str, Any]] = None,
    accept_cache: bool = True,
) -> Dict[str, Any]:
    headers = make_app_headers(api_key, token, accept_cache=accept_cache)
    return _request_json("POST", f"{base_url}{endpoint}", headers, json_body=body, params=params)


def taxpayer_post_api(
    base_url: str,
    endpoint: str,
    api_key: str,
    authorization_token: str,
    body: Dict[str, Any],
    params: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Used for Generate OTP and Verify OTP.
    authorization_token is the normal Sandbox access token until OTP is verified.
    """
    headers = make_taxpayer_headers(api_key, authorization_token)
    return _request_json("POST", f"{base_url}{endpoint}", headers, json_body=body, params=params)


def taxpayer_get_api(
    base_url: str,
    endpoint: str,
    api_key: str,
    taxpayer_token: str,
    params: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    headers = make_taxpayer_headers(api_key, taxpayer_token)
    return _request_json("GET", f"{base_url}{endpoint}", headers, params=params)


# ============================================================
# Public GSTIN APIs
# ============================================================
def find_first_profile_value(data: Any, candidate_keys: List[str]) -> str:
    """
    GST API providers sometimes change key names or nest turnover fields.
    This helper checks common candidate keys, first at root level and then recursively.
    """
    if not isinstance(data, (dict, list)):
        return ""

    candidates_lower = {key.lower(): key for key in candidate_keys}

    def scan(obj: Any) -> str:
        if isinstance(obj, dict):
            for key, value in obj.items():
                if str(key).lower() in candidates_lower and value not in [None, ""]:
                    if isinstance(value, (dict, list)):
                        return json.dumps(value, ensure_ascii=False)
                    return str(value)

            for value in obj.values():
                found = scan(value)
                if found:
                    return found

        elif isinstance(obj, list):
            for value in obj:
                found = scan(value)
                if found:
                    return found

        return ""

    return scan(data)


def find_keyword_paths(data: Any, keywords: List[str], max_items: int = 30) -> str:
    """
    Returns key paths containing keywords such as turnover, aato, gross, income, einvoice.
    Useful to confirm whether the API response actually contains turnover-related data.
    """
    matches: List[str] = []
    keywords_lower = [k.lower() for k in keywords]

    def walk(obj: Any, path: str = "") -> None:
        if len(matches) >= max_items:
            return

        if isinstance(obj, dict):
            for key, value in obj.items():
                key_text = str(key)
                new_path = f"{path}.{key_text}" if path else key_text
                if any(keyword in new_path.lower() for keyword in keywords_lower):
                    if isinstance(value, (dict, list)):
                        value_text = json.dumps(value, ensure_ascii=False)[:250]
                    else:
                        value_text = str(value)
                    matches.append(f"{new_path} = {value_text}")

                walk(value, new_path)

        elif isinstance(obj, list):
            for idx, value in enumerate(obj):
                walk(value, f"{path}[{idx}]")

    walk(data)
    return " | ".join(matches)


def extract_business_data(search_payload: Dict[str, Any]) -> Dict[str, Any]:
    data = get_nested(search_payload, "data", "data", default={})
    if not isinstance(data, dict):
        data = {}

    addr = get_nested(data, "pradr", "addr", default={})
    if not isinstance(addr, dict):
        addr = {}

    nba = data.get("nba", [])
    if isinstance(nba, list):
        nature_of_business = ", ".join([str(x) for x in nba])
    else:
        nature_of_business = str(nba or "")

    e_invoice_status = find_first_profile_value(
        data,
        [
            "einvoiceStatus",
            "eInvoiceStatus",
            "einvStatus",
            "einv_applicable",
            "isEinvoiceApplicable",
        ],
    )
    aggregate_turnover = find_first_profile_value(
        data,
        [
            "aggreTurnOver",
            "aggreTurnover",
            "aggre_turnover",
            "aggregateTurnover",
            "aggregate_turnover",
            "aggTurnOver",
            "aato",
            "AATO",
            "annualAggregateTurnover",
            "annual_aggregate_turnover",
        ],
    )
    aggregate_turnover_fy = find_first_profile_value(
        data,
        [
            "aggreTurnOverFY",
            "aggreTurnOverFy",
            "aggre_turnover_fy",
            "aggregateTurnoverFY",
            "aggregateTurnoverFy",
            "aggregate_turnover_fy",
            "aatoFinancialYear",
            "aatoFY",
            "AATOFY",
            "annualAggregateTurnoverFY",
            "annual_aggregate_turnover_fy",
        ],
    )
    gross_total_income = find_first_profile_value(
        data,
        [
            "grossTotalIncome",
            "gross_total_income",
            "gti",
            "GTI",
            "grossIncome",
            "gross_income",
        ],
    )
    gross_total_income_fy = find_first_profile_value(
        data,
        [
            "grossTotalIncomeFY",
            "grossTotalIncomeFy",
            "grossTotalIncomeFinancialYear",
            "gross_total_income_fy",
            "gtiFinancialYear",
            "gtiFY",
            "grossIncomeFY",
        ],
    )

    detected_keys = find_keyword_paths(
        data,
        ["turnover", "turn", "aato", "gross", "income", "gti", "einvoice", "e-invoice"],
    )
    turnover_found = any(
        str(value).strip()
        for value in [aggregate_turnover, aggregate_turnover_fy, gross_total_income, gross_total_income_fy]
    )
    turnover_status = "Found in API response" if turnover_found else "Not returned by this API response"

    return {
        "Legal Name of Business": data.get("lgnm", ""),
        "Trade Name": data.get("tradeNam", ""),
        "Constitution of Business": data.get("ctb", ""),
        "GSTIN / UIN Status": data.get("sts", ""),
        "Taxpayer Type": data.get("dty", ""),
        "Registration Date": data.get("rgdt", ""),
        "Cancellation Date": data.get("cxdt", ""),
        "Last Updated on GSTN": data.get("lstupdt", ""),
        "E-Invoice Status": e_invoice_status,
        "Aggregate Turnover": aggregate_turnover,
        "Aggregate Turnover FY": aggregate_turnover_fy,
        "Gross Total Income": gross_total_income,
        "Gross Total Income FY": gross_total_income_fy,
        "Turnover Data Status": turnover_status,
        "Detected Turnover/E-Invoice Keys": detected_keys,
        "Raw GSTIN Profile JSON": json.dumps(data, ensure_ascii=False)[:15000],
        "Nature of Business": nature_of_business,
        "Principal Place State": addr.get("stcd", ""),
        "Principal Place City": addr.get("loc", "") or addr.get("dst", ""),
        "Principal Place Pincode": addr.get("pncd", ""),
    }


def extract_preference(preference_payload: Dict[str, Any]) -> Tuple[str, str]:
    response = get_nested(preference_payload, "data", "data", "response", default=[])
    if not isinstance(response, list):
        return "", ""

    parts = []
    prefs = []
    for item in response:
        if not isinstance(item, dict):
            continue
        quarter = item.get("quarter", "")
        pref = item.get("preference", "")
        readable = {"M": "Monthly", "Q": "Quarterly"}.get(str(pref).upper(), str(pref))
        if readable:
            prefs.append(readable)
        if quarter or readable:
            parts.append(f"{quarter}: {readable}".strip(": "))

    if not parts:
        return "", ""

    unique = sorted(set(prefs))
    frequency = unique[0] if len(unique) == 1 else "Mixed"
    return frequency, "; ".join(parts)


def extract_filing_rows(track_payload: Dict[str, Any], gstin: str, financial_year: str) -> List[Dict[str, Any]]:
    rows = []
    filed_list = get_nested(track_payload, "data", "data", "EFiledlist", default=[])

    if not isinstance(filed_list, list):
        return rows

    for item in filed_list:
        if not isinstance(item, dict):
            continue

        rows.append(
            {
                "GSTIN": gstin,
                "Return Type": item.get("rtntype", ""),
                "Return Period": item.get("ret_prd", ""),
                "Date of Filing": item.get("dof", ""),
                "Status": item.get("status", ""),
                "ARN": item.get("arn", ""),
                "Mode of Filing": item.get("mof", ""),
                "Valid": item.get("valid", ""),
                "Financial Year": financial_year,
            }
        )

    return rows


def build_error(gstin: str, step: str, error: str, raw_response: Any = "") -> Dict[str, Any]:
    return {
        "GSTIN": gstin,
        "Step": step,
        "Category": classify_error(error),
        "Error": error,
        "Raw Response": str(raw_response)[:2000],
        "Fetched At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def blank_customer_row(gstin: str) -> Dict[str, Any]:
    row = {col: "" for col in CUSTOMER_COLUMNS}
    row["GSTIN"] = gstin
    row["Valid Format"] = False
    row["Checksum Valid"] = False
    row["API Status"] = "Not Started"
    row["Fetched At"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return row


def fetch_one_gstin(
    gstin: str,
    api_key: str,
    token: str,
    base_url: str,
    financial_year: str,
    fetch_profile: bool,
    fetch_filings: bool,
    fetch_preference: bool,
    accept_cache: bool,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]]]:
    gstin = clean_gstin(gstin)
    customer_row = blank_customer_row(gstin)
    customer_row["Valid Format"] = gstin_format_valid(gstin)
    customer_row["Checksum Valid"] = gstin_checksum_valid(gstin) if customer_row["Valid Format"] else False
    customer_row["State From GSTIN"] = gstin_state(gstin)
    customer_row["PAN From GSTIN"] = pan_from_gstin(gstin)

    filing_rows: List[Dict[str, Any]] = []
    error_rows: List[Dict[str, Any]] = []

    if not customer_row["Valid Format"]:
        customer_row["API Status"] = "Skipped"
        customer_row["API Message"] = "Invalid GSTIN format"
        return customer_row, filing_rows, [build_error(gstin, STEP_VALIDATION, "Invalid GSTIN format")]

    if not customer_row["Checksum Valid"]:
        customer_row["API Status"] = "Warning"
        customer_row["API Message"] = "GSTIN format is valid but checksum failed. API call skipped."
        return customer_row, filing_rows, [build_error(gstin, STEP_VALIDATION, "Checksum failed")]

    if fetch_profile:
        try:
            payload = post_api(
                base_url,
                SEARCH_GSTIN_ENDPOINT,
                api_key,
                token,
                {"gstin": gstin},
                accept_cache=accept_cache,
            )
            customer_row.update(extract_business_data(payload))
            customer_row["API Status"] = "Profile OK"
            customer_row["API Message"] = "GSTIN profile fetched"
        except Exception as exc:
            customer_row["API Status"] = "Profile Failed"
            customer_row["API Message"] = str(exc)
            error_rows.append(build_error(gstin, STEP_PROFILE, str(exc)))

    if fetch_preference:
        try:
            payload = post_api(
                base_url,
                PREFERENCE_ENDPOINT,
                api_key,
                token,
                {"gstin": gstin},
                params={"financial_year": financial_year},
                accept_cache=accept_cache,
            )
            frequency, preference_detail = extract_preference(payload)
            customer_row["Filing Frequency"] = frequency
            customer_row["Preference Detail"] = preference_detail
        except Exception as exc:
            if customer_row["API Status"] in ["Not Started", "Profile OK"]:
                customer_row["API Status"] = "Preference Failed"
            customer_row["API Message"] = (customer_row["API Message"] + " | " if customer_row["API Message"] else "") + str(exc)
            error_rows.append(build_error(gstin, STEP_PREFERENCE, str(exc)))

    if fetch_filings:
        try:
            payload = post_api(
                base_url,
                TRACK_GSTR_ENDPOINT,
                api_key,
                token,
                {"gstin": gstin},
                params={"financial_year": financial_year},
                accept_cache=accept_cache,
            )
            filing_rows = extract_filing_rows(payload, gstin, financial_year)
            if customer_row["API Status"] in ["Not Started", "Profile OK"]:
                customer_row["API Status"] = "OK"
            customer_row["API Message"] = (customer_row["API Message"] + " | " if customer_row["API Message"] else "") + f"{len(filing_rows)} filing rows fetched"
        except Exception as exc:
            if customer_row["API Status"] in ["Not Started", "Profile OK"]:
                customer_row["API Status"] = "Filing Failed"
            customer_row["API Message"] = (customer_row["API Message"] + " | " if customer_row["API Message"] else "") + str(exc)
            error_rows.append(build_error(gstin, STEP_FILINGS, str(exc)))

    if customer_row["API Status"] == "Not Started":
        customer_row["API Status"] = "Completed"
        customer_row["API Message"] = "No API option selected"

    return customer_row, filing_rows, error_rows


# ============================================================
# Bulk run orchestration (incremental, optionally parallel)
# ============================================================
def partition_targets(
    gstin_list: List[str],
    customers_df: pd.DataFrame,
    force_refresh: bool,
) -> Tuple[List[str], List[str]]:
    """Splits GSTINs into ones that still need fetching and ones already successfully fetched."""
    if force_refresh or customers_df.empty:
        return list(gstin_list), []

    done = set(customers_df.loc[customers_df[COL_API_STATUS].isin(SUCCESS_API_STATUSES), COL_GSTIN])
    to_fetch = [g for g in gstin_list if g not in done]
    cached = [g for g in gstin_list if g in done]
    return to_fetch, cached


def merge_run_results(
    customer_rows: List[Dict[str, Any]],
    filing_rows: List[Dict[str, Any]],
    error_rows: List[Dict[str, Any]],
    affected_gstins: List[str],
    attempted_steps: List[str],
) -> None:
    """Upserts one bulk-run's results into session_state, by GSTIN (and by step for errors)."""
    new_customers = pd.DataFrame(customer_rows, columns=CUSTOMER_COLUMNS)
    new_filings = pd.DataFrame(filing_rows, columns=FILING_COLUMNS)
    new_errors = pd.DataFrame(error_rows, columns=ERROR_COLUMNS)

    customers = st.session_state["customers_df"].reindex(columns=CUSTOMER_COLUMNS)
    customers = customers[~customers[COL_GSTIN].isin(affected_gstins)]
    st.session_state["customers_df"] = pd.concat([customers, new_customers], ignore_index=True).reindex(columns=CUSTOMER_COLUMNS)

    if STEP_FILINGS in attempted_steps:
        filings = st.session_state["filings_df"].reindex(columns=FILING_COLUMNS)
        filings = filings[~filings[COL_GSTIN].isin(affected_gstins)]
        st.session_state["filings_df"] = pd.concat([filings, new_filings], ignore_index=True).reindex(columns=FILING_COLUMNS)

    errors = st.session_state["errors_df"].reindex(columns=ERROR_COLUMNS)
    if affected_gstins and attempted_steps:
        drop_mask = errors[COL_GSTIN].isin(affected_gstins) & errors["Step"].isin(attempted_steps)
        errors = errors[~drop_mask]
    st.session_state["errors_df"] = pd.concat([errors, new_errors], ignore_index=True).reindex(columns=ERROR_COLUMNS)


def execute_bulk_run(
    targets: List[str],
    api_key: str,
    api_secret: str,
    base_url: str,
    financial_year: str,
    fetch_profile: bool,
    fetch_filings: bool,
    fetch_preference: bool,
    accept_cache: bool,
    max_workers: int,
    delay_seconds: float,
) -> None:
    if not targets:
        st.info("Nothing to process — every selected GSTIN already has a result. Turn on Force refresh to re-fetch.")
        return

    attempted_steps = [STEP_VALIDATION]
    if fetch_profile:
        attempted_steps.append(STEP_PROFILE)
    if fetch_preference:
        attempted_steps.append(STEP_PREFERENCE)
    if fetch_filings:
        attempted_steps.append(STEP_FILINGS)

    customer_rows: List[Dict[str, Any]] = []
    filing_rows: List[Dict[str, Any]] = []
    error_rows: List[Dict[str, Any]] = []

    with st.status(f"Running bulk lookup for {len(targets)} GSTIN(s)…", expanded=True) as status_box:
        try:
            token = authenticate(api_key, api_secret, base_url)
        except Exception as exc:
            set_connection_status(False, str(exc))
            log_activity("Authentication failed", str(exc), level="error")
            status_box.update(label="Authentication failed", state="error")
            st.error(f"Authentication failed: {exc}")
            return

        set_connection_status(True, "Authenticated for bulk run")
        status_box.write("✅ Authenticated with the API provider.")

        progress_bar = st.progress(0.0)
        total = len(targets)
        completed = 0

        def _fetch(gstin: str):
            return fetch_one_gstin(
                gstin=gstin,
                api_key=api_key,
                token=token,
                base_url=base_url,
                financial_year=financial_year,
                fetch_profile=fetch_profile,
                fetch_filings=fetch_filings,
                fetch_preference=fetch_preference,
                accept_cache=accept_cache,
            )

        def _record(gstin: str, cust: Dict[str, Any], fil: List[Dict[str, Any]], err: List[Dict[str, Any]]) -> None:
            nonlocal completed
            customer_rows.append(cust)
            filing_rows.extend(fil)
            error_rows.extend(err)
            completed += 1
            progress_bar.progress(completed / total)
            mark = "✅" if cust.get("API Status") not in FAILURE_API_STATUSES else "❌"
            status_box.write(f"{mark} `{gstin}` — {cust.get('API Status', '')}")

        if max_workers <= 1:
            for gstin in targets:
                cust, fil, err = _fetch(gstin)
                _record(gstin, cust, fil, err)
                if delay_seconds:
                    time.sleep(float(delay_seconds))
        else:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_map = {executor.submit(_fetch, g): g for g in targets}
                for future in as_completed(future_map):
                    gstin = future_map[future]
                    try:
                        cust, fil, err = future.result()
                    except Exception as exc:
                        cust = blank_customer_row(gstin)
                        cust["API Status"] = "Failed"
                        cust["API Message"] = str(exc)
                        fil, err = [], [build_error(gstin, "Unexpected", str(exc))]
                    _record(gstin, cust, fil, err)

        status_box.update(label=f"Bulk lookup completed — {len(targets)} processed", state="complete")

    merge_run_results(customer_rows, filing_rows, error_rows, targets, attempted_steps)

    failed_count = sum(1 for c in customer_rows if c.get("API Status") in FAILURE_API_STATUSES)
    log_activity(
        "Bulk lookup run",
        f"{len(targets)} GSTIN(s) processed, {failed_count} failed.",
        level="warning" if failed_count else "success",
    )
    st.toast(
        f"Bulk lookup finished — {len(targets) - failed_count} ok, {failed_count} failed.",
        icon="✅" if not failed_count else "⚠️",
    )


# ============================================================
# Taxpayer Authentication + Private Return APIs
# ============================================================
def generate_taxpayer_otp(
    base_url: str,
    api_key: str,
    app_token: str,
    gstin: str,
    username: str,
) -> Dict[str, Any]:
    return taxpayer_post_api(
        base_url,
        TAXPAYER_OTP_ENDPOINT,
        api_key,
        app_token,
        {"username": username.strip(), "gstin": clean_gstin(gstin)},
    )


def verify_taxpayer_otp(
    base_url: str,
    api_key: str,
    app_token: str,
    gstin: str,
    username: str,
    otp: str,
) -> Dict[str, Any]:
    return taxpayer_post_api(
        base_url,
        TAXPAYER_OTP_VERIFY_ENDPOINT,
        api_key,
        app_token,
        {"username": username.strip(), "gstin": clean_gstin(gstin)},
        params={"otp": otp.strip()},
    )


def fetch_gstr3b_details(
    base_url: str,
    api_key: str,
    taxpayer_token: str,
    year: str,
    month: str,
) -> Dict[str, Any]:
    endpoint = GSTR3B_DETAILS_ENDPOINT_TEMPLATE.format(year=year.strip(), month=month.strip())
    return taxpayer_get_api(base_url, endpoint, api_key, taxpayer_token)


def fetch_gstr1_section(
    base_url: str,
    api_key: str,
    taxpayer_token: str,
    section_slug: str,
    year: str,
    month: str,
) -> Dict[str, Any]:
    endpoint = f"/gst/compliance/tax-payer/gstrs/gstr-1/{section_slug}/{year.strip()}/{month.strip()}"
    return taxpayer_get_api(base_url, endpoint, api_key, taxpayer_token)


def response_status_message(payload: Dict[str, Any]) -> Tuple[str, str]:
    status_cd = get_nested(payload, "data", "status_cd", default="")
    err = get_nested(payload, "data", "error", default={})
    if isinstance(err, dict) and err:
        return str(status_cd or "0"), f"{err.get('error_cd', '')} {err.get('message', '')}".strip()
    if str(status_cd) == "1":
        return "1", "Success"
    return str(status_cd or ""), ""


def json_leaf_rows(
    obj: Any,
    context: Dict[str, Any],
    path: str = "",
    rows: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    if rows is None:
        rows = []

    if isinstance(obj, dict):
        for key, value in obj.items():
            new_path = f"{path}.{key}" if path else str(key)
            json_leaf_rows(value, context, new_path, rows)
    elif isinstance(obj, list):
        for idx, value in enumerate(obj):
            new_path = f"{path}[{idx}]"
            json_leaf_rows(value, context, new_path, rows)
    else:
        row = dict(context)
        row["Path"] = path
        row["Value"] = obj
        rows.append(row)

    return rows


def sum_item_amounts(invoice: Dict[str, Any]) -> Dict[str, float]:
    totals = {"txval": 0.0, "iamt": 0.0, "camt": 0.0, "samt": 0.0, "csamt": 0.0}
    itms = invoice.get("itms", [])
    if not isinstance(itms, list):
        return totals

    for item in itms:
        if not isinstance(item, dict):
            continue
        itm_det = item.get("itm_det", {})
        if not isinstance(itm_det, dict):
            continue
        for key in totals:
            try:
                totals[key] += float(itm_det.get(key, 0) or 0)
            except Exception:
                pass

    return totals


def find_invoice_like_dicts(
    obj: Any,
    section: str,
    ancestors: Optional[List[Dict[str, Any]]] = None,
    path: str = "",
    out: Optional[List[Tuple[str, Dict[str, Any], List[Dict[str, Any]]]]] = None,
) -> List[Tuple[str, Dict[str, Any], List[Dict[str, Any]]]]:
    """
    Finds invoice/note-like dictionaries in GSTR-1 responses.
    This makes the app useful even when different sections have slightly different schemas.
    """
    if ancestors is None:
        ancestors = []
    if out is None:
        out = []

    if isinstance(obj, dict):
        keys = set(obj.keys())
        invoice_markers = {"inum", "idt", "val", "inv_typ"}
        note_markers = {"nt_num", "nt_dt", "ntty"}
        doc_markers = {"doc_num", "from", "to", "totnum", "net_issue"}

        is_invoice = bool(invoice_markers.intersection(keys)) and ("itms" in keys or "val" in keys or "inum" in keys)
        is_note = bool(note_markers.intersection(keys))
        is_doc = section == "DOC-ISSUE" and bool(doc_markers.intersection(keys))

        if is_invoice or is_note or is_doc:
            out.append((path, obj, ancestors.copy()))

        new_ancestors = ancestors + [obj]
        for key, value in obj.items():
            find_invoice_like_dicts(value, section, new_ancestors, f"{path}.{key}" if path else str(key), out)

    elif isinstance(obj, list):
        for idx, item in enumerate(obj):
            find_invoice_like_dicts(item, section, ancestors, f"{path}[{idx}]", out)

    return out


def extract_gstr1_invoice_rows(payload: Dict[str, Any], gstin: str, year: str, month: str, section: str) -> List[Dict[str, Any]]:
    data = get_nested(payload, "data", "data", default={})
    rows = []
    found = find_invoice_like_dicts(data, section)

    for path, doc, ancestors in found:
        counterparty = ""
        for anc in reversed(ancestors):
            if isinstance(anc, dict) and anc.get("ctin"):
                counterparty = str(anc.get("ctin"))
                break

        totals = sum_item_amounts(doc)
        rows.append(
            {
                "GSTIN": gstin,
                "Year": year,
                "Month": month,
                "Section": section,
                "Counterparty GSTIN": counterparty,
                "Invoice Number": doc.get("inum") or doc.get("nt_num") or doc.get("doc_num") or "",
                "Invoice Date": doc.get("idt") or doc.get("nt_dt") or "",
                "Invoice Value": doc.get("val") or "",
                "Invoice Type": doc.get("inv_typ") or doc.get("ntty") or "",
                "POS": doc.get("pos") or "",
                "Reverse Charge": doc.get("rchrg") or "",
                "Taxable Value": totals["txval"],
                "IGST": totals["iamt"],
                "CGST": totals["camt"],
                "SGST": totals["samt"],
                "CESS": totals["csamt"],
                "Source Type": doc.get("srctyp") or "",
                "Raw Path": path,
            }
        )

    return rows


def add_return_response(
    gstin: str,
    ret: str,
    section: str,
    year: str,
    month: str,
    payload: Dict[str, Any],
) -> Dict[str, Any]:
    status, msg = response_status_message(payload)
    return {
        "GSTIN": gstin,
        "Return": ret,
        "Section": section,
        "Year": year,
        "Month": month,
        "Status": status,
        "Message": msg,
        "Transaction ID": payload.get("transaction_id", ""),
        "Fetched At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Raw JSON": json.dumps(payload, ensure_ascii=False),
    }


def make_return_facts(
    gstin: str,
    ret: str,
    section: str,
    year: str,
    month: str,
    payload: Dict[str, Any],
) -> pd.DataFrame:
    data = get_nested(payload, "data", "data", default=payload)
    context = {
        "GSTIN": gstin,
        "Return": ret,
        "Section": section,
        "Year": year,
        "Month": month,
    }
    rows = json_leaf_rows(data, context)
    return pd.DataFrame(rows, columns=RETURN_FACT_COLUMNS)


# ============================================================
# Input / Filter / Display helpers
# ============================================================
def read_uploaded_gstins(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()

    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file, dtype=str)
    if name.endswith((".xlsx", ".xls")):
        return pd.read_excel(uploaded_file, dtype=str)

    raise ValueError("Upload only CSV or Excel.")


def sample_template() -> bytes:
    df = pd.DataFrame({"GSTIN": ["27ABCDE1234F1Z5", "07ABCDE1234F1Z2"]})
    return df.to_csv(index=False).encode("utf-8")


def filter_dataframe(df: pd.DataFrame, query: str, columns: Optional[List[str]] = None) -> pd.DataFrame:
    if not query or df.empty:
        return df
    cols = [c for c in (columns or list(df.columns)) if c in df.columns]
    if not cols:
        return df
    mask = pd.Series(False, index=df.index)
    needle = query.strip()
    for col in cols:
        mask = mask | df[col].astype(str).str.contains(needle, case=False, na=False, regex=False)
    return df[mask]


def badge(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    lowered = text.lower()
    if re.search(r"\bok\b", lowered):
        return f"✅ {text}"
    for token, icon in STATUS_ICON_RULES:
        if token in lowered:
            return f"{icon} {text}"
    return f"⚪ {text}"


def decorate_for_display(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    """Returns a COPY with status columns badge-prefixed for on-screen display only (exports stay clean)."""
    out = df.copy()
    for col in columns:
        if col in out.columns:
            out[col] = out[col].map(badge)
    return out


def html_badge(text: str, kind: str = "idle") -> str:
    cls = {"ok": "badge-ok", "warn": "badge-warn", "bad": "badge-bad"}.get(kind, "badge-idle")
    return f'<span class="badge {cls}">{html.escape(str(text))}</span>'


def render_gstin_anatomy(gstin: str) -> str:
    g = clean_gstin(gstin)
    seg = gstin_segments(g)

    def cell(css_class: str, value: str, label: str) -> str:
        safe_value = html.escape(value) if value else "·"
        return f'<div class="gstin-seg {css_class}"><span class="val">{safe_value}</span><span class="lbl">{label}</span></div>'

    parts = [
        cell("seg-state", seg["state"], "State"),
        cell("seg-pan", seg["pan"], "PAN"),
        cell("seg-entity", seg["entity"], "Entity"),
        cell("seg-z", seg["z"], "Default"),
        cell("seg-check", seg["checksum"], "Check"),
    ]
    return f'<div class="gstin-anatomy">{"".join(parts)}</div>'


def render_status_pie(df: pd.DataFrame, column: str, title: str) -> None:
    if df.empty or column not in df.columns:
        st.caption(f"No data yet for {title.lower()}.")
        return
    series = df[column].astype(str).str.strip().replace("", "Unknown")
    counts = series.value_counts().reset_index()
    counts.columns = [column, "Count"]
    if counts.empty:
        st.caption(f"No data yet for {title.lower()}.")
        return
    fig = px.pie(counts, names=column, values="Count", hole=0.55, color_discrete_sequence=CHART_PALETTE)
    fig.update_traces(textinfo="percent+label", textfont_size=11)
    fig.update_layout(
        title=title,
        margin=dict(t=42, b=10, l=10, r=10),
        showlegend=False,
        height=300,
        font=dict(family="Inter, sans-serif"),
    )
    st.plotly_chart(fig, use_container_width=True)


def render_state_bar(df: pd.DataFrame, column: str, title: str) -> None:
    if df.empty or column not in df.columns:
        st.caption(f"No data yet for {title.lower()}.")
        return
    series = df[column].astype(str).str.strip().replace("", "Unknown")
    counts = series.value_counts().reset_index()
    counts.columns = [column, "Count"]
    counts = counts.sort_values("Count", ascending=True).tail(12)
    if counts.empty:
        st.caption(f"No data yet for {title.lower()}.")
        return
    fig = px.bar(counts, x="Count", y=column, orientation="h", color_discrete_sequence=[CHART_PALETTE[0]])
    fig.update_layout(
        title=title,
        margin=dict(t=42, b=10, l=10, r=10),
        height=300,
        font=dict(family="Inter, sans-serif"),
        yaxis_title="",
        xaxis_title="",
    )
    st.plotly_chart(fig, use_container_width=True)


# ============================================================
# Export helpers
# ============================================================
def autosize_and_style(ws, df: pd.DataFrame, header_color: str = "1E3A8A") -> None:
    if len(df.columns) == 0:
        return
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        sample = df[column].astype(str).head(300)
        longest = max([len(str(column))] + [len(v) for v in sample]) if len(sample) else len(str(column))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(longest + 2, 10), 60)
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions


def colorize_status_column(ws, df: pd.DataFrame, column_name: str) -> None:
    if column_name not in df.columns or df.empty:
        return
    col_idx = list(df.columns).index(column_name) + 1
    fill_ok = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_warn = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_bad = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, value in enumerate(df[column_name].astype(str), start=2):
        lowered = value.lower()
        if "cancel" in lowered or "fail" in lowered or "inactive" in lowered:
            fill = fill_bad
        elif "warning" in lowered or "skip" in lowered or "suspend" in lowered:
            fill = fill_warn
        elif "active" in lowered or "ok" in lowered or "complet" in lowered:
            fill = fill_ok
        else:
            continue
        ws.cell(row=row_idx, column=col_idx).fill = fill


def make_excel(
    customers: pd.DataFrame,
    filings: pd.DataFrame,
    errors: pd.DataFrame,
    return_responses: Optional[pd.DataFrame] = None,
    return_facts: Optional[pd.DataFrame] = None,
    gstr1_invoices: Optional[pd.DataFrame] = None,
) -> bytes:
    output = BytesIO()

    return_responses = return_responses if return_responses is not None else pd.DataFrame(columns=RETURN_RESPONSE_COLUMNS)
    return_facts = return_facts if return_facts is not None else pd.DataFrame(columns=RETURN_FACT_COLUMNS)
    gstr1_invoices = gstr1_invoices if gstr1_invoices is not None else pd.DataFrame(columns=GSTR1_INVOICE_COLUMNS)

    active_count = int(customers[COL_GSTIN_STATUS].astype(str).str.contains("Active", case=False, na=False).sum()) if not customers.empty else 0
    ok_count = int(customers[COL_API_STATUS].astype(str).str.contains("OK|Profile OK|Completed", case=False, na=False).sum()) if not customers.empty else 0

    summary = pd.DataFrame(
        [
            {"Metric": "Report Generated At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"Metric": "Total GSTINs", "Value": len(customers)},
            {"Metric": "Successful / Partial Rows", "Value": ok_count},
            {"Metric": "Active GSTINs", "Value": active_count},
            {"Metric": "Public Filing Rows", "Value": len(filings)},
            {"Metric": "Private Return Responses", "Value": len(return_responses)},
            {"Metric": "GSTR-1 Invoice Rows", "Value": len(gstr1_invoices)},
            {"Metric": "Return Fact Rows", "Value": len(return_facts)},
            {"Metric": "Error Rows", "Value": len(errors)},
        ]
    )

    sheets: List[Tuple[str, pd.DataFrame]] = [
        ("Summary", summary),
        ("Taxpayer Details", customers),
        ("Public Filing Table", filings),
        ("GSTR1 Invoice Rows", gstr1_invoices),
        ("Return Facts", return_facts),
        ("Raw Return JSON", return_responses),
        ("Errors", errors),
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, frame in sheets:
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
            autosize_and_style(writer.sheets[sheet_name], frame)

        colorize_status_column(writer.sheets["Taxpayer Details"], customers, COL_API_STATUS)
        colorize_status_column(writer.sheets["Taxpayer Details"], customers, COL_GSTIN_STATUS)

    return output.getvalue()


def make_json_export(
    customers: pd.DataFrame,
    filings: pd.DataFrame,
    errors: pd.DataFrame,
    return_facts: pd.DataFrame,
    gstr1_invoices: pd.DataFrame,
) -> bytes:
    payload = {
        "generated_at": datetime.now().isoformat(),
        "taxpayer_details": json.loads(customers.to_json(orient="records")),
        "public_filing_table": json.loads(filings.to_json(orient="records")),
        "gstr1_invoice_rows": json.loads(gstr1_invoices.to_json(orient="records")),
        "return_facts": json.loads(return_facts.to_json(orient="records")),
        "errors": json.loads(errors.to_json(orient="records")),
    }
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


def append_df(existing: pd.DataFrame, incoming: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    if incoming.empty:
        return existing.reindex(columns=columns)
    combined = pd.concat([existing.reindex(columns=columns), incoming.reindex(columns=columns)], ignore_index=True)
    return combined.reindex(columns=columns)


# ============================================================
# Session state, logging, reset
# ============================================================
def init_state() -> None:
    defaults: Dict[str, Any] = {
        "gstin_list": [],
        "customers_df": pd.DataFrame(columns=CUSTOMER_COLUMNS),
        "filings_df": pd.DataFrame(columns=FILING_COLUMNS),
        "errors_df": pd.DataFrame(columns=ERROR_COLUMNS),
        "taxpayer_sessions": {},
        "pending_taxpayer_auth": {},
        "return_responses_df": pd.DataFrame(columns=RETURN_RESPONSE_COLUMNS),
        "return_facts_df": pd.DataFrame(columns=RETURN_FACT_COLUMNS),
        "gstr1_invoices_df": pd.DataFrame(columns=GSTR1_INVOICE_COLUMNS),
        "activity_log": [],
        "connection_status": {"ok": None, "message": "Not checked yet", "checked_at": ""},
        "gstr1_section_pick": list(COMMON_GSTR1_SECTIONS),
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def reset_all_data() -> None:
    for key in RESETTABLE_KEYS:
        if key in st.session_state:
            del st.session_state[key]
    init_state()


def set_connection_status(ok: bool, message: str) -> None:
    st.session_state["connection_status"] = {
        "ok": ok,
        "message": message,
        "checked_at": datetime.now().strftime("%H:%M:%S"),
    }


def log_activity(action: str, detail: str = "", level: str = "info") -> None:
    entry = {
        "Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Level": level,
        "Action": action,
        "Detail": detail,
    }
    st.session_state["activity_log"].insert(0, entry)
    st.session_state["activity_log"] = st.session_state["activity_log"][:300]


def render_connection_badge() -> None:
    status = st.session_state.get("connection_status", {})
    ok = status.get("ok")
    checked_at = status.get("checked_at", "")
    suffix = f" · {checked_at}" if checked_at else ""
    if ok is True:
        st.markdown(f'<span class="chip"><span class="dot dot-ok"></span><b>Connected</b>{suffix}</span>', unsafe_allow_html=True)
    elif ok is False:
        st.markdown(f'<span class="chip"><span class="dot dot-bad"></span><b>Connection failed</b>{suffix}</span>', unsafe_allow_html=True)
    else:
        st.markdown('<span class="chip"><span class="dot dot-idle"></span>Not verified yet</span>', unsafe_allow_html=True)


# ============================================================
# Streamlit UI
# ============================================================
init_state()

st.set_page_config(
    page_title="GSTIN Compliance Suite",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)

BASE_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600&family=IBM+Plex+Mono:wght@500;600&display=swap');

:root {
    --bg: #F3F5F9;
    --surface: #FFFFFF;
    --ink-900: #0E1428;
    --ink-600: #3D4866;
    --ink-400: #7C879F;
    --line: #E2E6EE;
    --brand-indigo: #1E3A8A;
    --brand-teal: #0F766E;
    --ok: #15803D;
    --warn: #B45309;
    --bad: #B91C1C;
    --neutral: #64748B;
    --accent: #C2730A;
    --accent-soft: #FCEEDC;
    --radius: 12px;
}

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

.stApp {
    background: var(--bg);
}

.block-container {
    padding-top: 1.1rem;
    padding-bottom: 3rem;
    max-width: 1240px;
}

h1, h2, h3, h4 {
    font-family: 'Space Grotesk', sans-serif;
    color: var(--ink-900);
    letter-spacing: -0.01em;
}

code, .mono {
    font-family: 'IBM Plex Mono', monospace !important;
}

.letterhead {
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 0.75rem;
    padding: 1rem 1.4rem;
    background: var(--surface);
    border: 1px solid var(--line);
    border-bottom: 3px solid var(--accent);
    border-radius: var(--radius);
    margin-bottom: 1.1rem;
}
.letterhead-title {
    display: flex;
    align-items: baseline;
    gap: 0.6rem;
    flex-wrap: wrap;
}
.letterhead-title h1 {
    font-size: 1.5rem;
    margin: 0;
    font-weight: 700;
}
.letterhead-title .tag {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.72rem;
    color: var(--ink-400);
    border: 1px solid var(--line);
    padding: 0.1rem 0.45rem;
    border-radius: 6px;
}
.letterhead-sub {
    font-size: 0.86rem;
    color: var(--ink-400);
    margin-top: 0.15rem;
}
.chip-row {
    display: flex;
    gap: 0.5rem;
    flex-wrap: wrap;
}
.chip {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.74rem;
    padding: 0.32rem 0.62rem;
    border-radius: 999px;
    border: 1px solid var(--line);
    background: var(--bg);
    color: var(--ink-600);
    white-space: nowrap;
}
.chip b { color: var(--ink-900); }
.dot { display: inline-block; width: 7px; height: 7px; border-radius: 50%; margin-right: 5px; }
.dot-ok { background: var(--ok); }
.dot-bad { background: var(--bad); }
.dot-idle { background: var(--neutral); }

.eyebrow {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    color: var(--accent);
    margin-bottom: 0.3rem;
    display: block;
}

.gstin-anatomy {
    display: flex;
    gap: 0.35rem;
    flex-wrap: wrap;
    align-items: flex-end;
    margin: 0.6rem 0 0.2rem 0;
}
.gstin-seg {
    display: flex;
    flex-direction: column;
    align-items: center;
}
.gstin-seg .val {
    font-family: 'IBM Plex Mono', monospace;
    font-weight: 600;
    font-size: 1.0rem;
    padding: 0.3rem 0.5rem;
    border-radius: 8px;
    min-width: 1.3rem;
    text-align: center;
    border: 1px solid var(--line);
}
.gstin-seg .lbl {
    font-size: 0.62rem;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    color: var(--ink-400);
    margin-top: 0.25rem;
}
.seg-state .val { background: #E4E9F8; color: var(--brand-indigo); }
.seg-pan .val { background: #DFF3F1; color: var(--brand-teal); }
.seg-entity .val { background: #F1F5F9; color: var(--ink-600); }
.seg-z .val { background: #F1F5F9; color: var(--ink-600); }
.seg-check .val { background: var(--accent-soft); color: var(--accent); }

.badge {
    display: inline-block;
    padding: 0.15rem 0.55rem;
    border-radius: 999px;
    font-size: 0.78rem;
    font-weight: 600;
}
.badge-ok { background: #DCFCE7; color: var(--ok); }
.badge-warn { background: #FEF3C7; color: var(--warn); }
.badge-bad { background: #FEE2E2; color: var(--bad); }
.badge-idle { background: #F1F5F9; color: var(--neutral); }

.log-list { display: flex; flex-direction: column; }
.log-row { display: flex; gap: 0.7rem; padding: 0.6rem 0; border-bottom: 1px solid var(--line); align-items: flex-start; }
.log-row .badge { min-width: 5.6rem; text-align: center; }
.log-body { flex: 1; }
.log-body b { color: var(--ink-900); font-size: 0.92rem; }
.log-detail { font-size: 0.83rem; color: var(--ink-400); margin-top: 0.1rem; }
.log-time { font-size: 0.78rem; color: var(--ink-400); white-space: nowrap; }

div[data-testid="stMetric"] {
    background: var(--surface);
    border: 1px solid var(--line);
    padding: 0.85rem 1rem;
    border-radius: var(--radius);
}
div[data-testid="stMetricLabel"] { color: var(--ink-400); }
div[data-testid="stMetricValue"] { font-family: 'Space Grotesk', sans-serif; color: var(--ink-900); }

div[data-testid="stTabs"] button {
    font-family: 'Space Grotesk', sans-serif;
    font-weight: 600;
    font-size: 0.92rem;
    color: var(--ink-400);
}
div[data-testid="stTabs"] [aria-selected="true"] {
    color: var(--ink-900) !important;
    border-bottom: 2.5px solid var(--accent) !important;
}
div[data-testid="stTabs"] [data-baseweb="tab-list"] {
    border-bottom: 1px solid var(--line);
    gap: 0.4rem;
}

section[data-testid="stSidebar"] {
    background: var(--surface);
    border-right: 1px solid var(--line);
}

div[data-testid="stExpander"], div[data-testid="stStatusWidget"] {
    border: 1px solid var(--line);
    border-radius: var(--radius);
}

div[data-testid="stDataFrame"] {
    border-radius: var(--radius);
    overflow: hidden;
    border: 1px solid var(--line);
}

div[data-testid="stButton"] button[kind="primary"],
div[data-testid="stDownloadButton"] button[kind="primary"] {
    background: var(--ink-900);
    border: none;
}
div[data-testid="stButton"] button[kind="primary"]:hover,
div[data-testid="stDownloadButton"] button[kind="primary"]:hover {
    background: var(--accent);
}

footer { visibility: hidden; }
.app-footer {
    text-align: center;
    color: var(--ink-400);
    font-size: 0.78rem;
    padding: 1.4rem 0 0.5rem 0;
    border-top: 1px solid var(--line);
    margin-top: 2rem;
}
</style>
"""
st.markdown(BASE_CSS, unsafe_allow_html=True)

# ------------------------------------------------------------
# Sidebar — Control Panel
# ------------------------------------------------------------
with st.sidebar:
    st.markdown("#### ⚙️ Control Panel")
    st.caption("Configure your connection, lookup behaviour, and appearance.")

    with st.expander("🔌 Connection", expanded=True):
        env_default = os.getenv("SANDBOX_ENV", "production").lower()
        env = st.selectbox("Environment", ["production", "test"], index=0 if env_default == "production" else 1)
        base_url = get_base_url(env)

        api_key = st.text_input("Sandbox API Key", value=os.getenv("SANDBOX_API_KEY", ""), type="password")
        api_secret = st.text_input("Sandbox API Secret", value=os.getenv("SANDBOX_API_SECRET", ""), type="password")

        render_connection_badge()

        if st.button("Test connection", use_container_width=True):
            try:
                authenticate(api_key, api_secret, base_url)
                set_connection_status(True, "Authentication successful")
                log_activity("Connection test", "Authentication succeeded", level="success")
                st.rerun()
            except Exception as exc:
                set_connection_status(False, str(exc))
                log_activity("Connection test", str(exc), level="error")
                st.error(f"Authentication failed: {exc}")

    with st.expander("📋 Lookup Options", expanded=True):
        financial_year = st.text_input("Financial year for public filing status", value="FY 2026-27", help="Example: FY 2026-27")
        fetch_profile = st.checkbox("Fetch taxpayer profile", value=True)
        fetch_filings = st.checkbox("Fetch public filing table", value=True)
        fetch_preference = st.checkbox("Fetch filing frequency", value=True)
        accept_cache = st.checkbox("Accept cached API response", value=True, help="Uses the provider's x-accept-cache header to speed up repeat lookups.")

    with st.expander("🚦 Performance & Reliability", expanded=False):
        max_workers = st.slider("Parallel workers", min_value=1, max_value=8, value=1, help="Run more than 1 to fetch several GSTINs at once. Keep this modest to avoid hitting API rate limits.")
        delay_seconds = st.number_input("Delay between calls (seconds)", min_value=0.0, max_value=10.0, value=0.3, step=0.1, help="Only applied when Parallel workers = 1.")
        force_refresh = st.checkbox("Force refresh already-fetched GSTINs", value=False, help="By default, GSTINs that already have a successful result are skipped on the next run to save API calls.")

    with st.expander("🎨 Appearance", expanded=False):
        accent_name = st.selectbox("Accent colour", list(ACCENTS.keys()), index=0)

    st.download_button(
        "⬇️ Download CSV template",
        data=sample_template(),
        file_name="bulk_gstin_template.csv",
        mime="text/csv",
        use_container_width=True,
    )

    st.divider()
    with st.popover("🧹 Reset all data", use_container_width=True):
        st.write("This clears every fetched GSTIN, taxpayer session, and log from this run. Your API credentials and settings stay as they are.")
        if st.button("Yes, clear everything", type="primary"):
            reset_all_data()
            log_activity("Reset", "All fetched data cleared", level="warning")
            st.rerun()

theme = ACCENTS[accent_name]
st.markdown(
    f"<style>:root {{ --accent: {theme['accent']}; --accent-soft: {theme['accent_soft']}; }}</style>",
    unsafe_allow_html=True,
)

# ------------------------------------------------------------
# Letterhead
# ------------------------------------------------------------
gstin_count = len(st.session_state.get("gstin_list", []))
processed_count = len(st.session_state.get("customers_df", pd.DataFrame()))
active_sessions = len(st.session_state.get("taxpayer_sessions", {}))

st.markdown(
    f"""
    <div class="letterhead">
        <div>
            <div class="letterhead-title">
                <h1>🧾 GSTIN Compliance Suite</h1>
                <span class="tag">sandbox.co.in API</span>
            </div>
            <div class="letterhead-sub">Bulk GSTIN profiles, filing tracker, and OTP-authorised GSTR‑1 / GSTR‑3B data, in one place.</div>
        </div>
        <div class="chip-row">
            <span class="chip">Environment · <b>{env.title()}</b></span>
            <span class="chip">GSTINs loaded · <b>{gstin_count}</b></span>
            <span class="chip">Processed · <b>{processed_count}</b></span>
            <span class="chip">Active sessions · <b>{active_sessions}</b></span>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ------------------------------------------------------------
# Quick GSTIN checker
# ------------------------------------------------------------
with st.container(border=True):
    st.markdown('<span class="eyebrow">Quick check</span>', unsafe_allow_html=True)
    qc1, qc2 = st.columns([1, 2])
    with qc1:
        quick_gstin = st.text_input("Check a single GSTIN", placeholder="e.g. 27ABCDE1234F1Z5", label_visibility="collapsed")
        st.caption("No API call — instant offline check.")
    with qc2:
        if quick_gstin.strip():
            cleaned = clean_gstin(quick_gstin)
            fmt_ok = gstin_format_valid(cleaned)
            chk_ok = gstin_checksum_valid(cleaned) if fmt_ok else False
            state_name = gstin_state(cleaned)
            pan = pan_from_gstin(cleaned)
            line1 = html_badge("Valid format", "ok") if fmt_ok else html_badge("Invalid format", "bad")
            line2 = html_badge("Checksum OK", "ok") if chk_ok else html_badge("Checksum mismatch", "bad")
            st.markdown(
                f'{line1} &nbsp; {line2} &nbsp; <span class="chip">{html.escape(state_name)}</span> '
                f'&nbsp; <span class="chip mono">PAN {html.escape(pan) if pan else "—"}</span>',
                unsafe_allow_html=True,
            )
            st.markdown(render_gstin_anatomy(cleaned), unsafe_allow_html=True)
        else:
            st.caption("Type or paste a GSTIN above to see its format, checksum, state, and embedded PAN.")

with st.expander("What credentials are needed?", expanded=not bool(api_key and api_secret)):
    cred1, cred2 = st.columns(2)
    with cred1:
        st.markdown("**Public data — already works today**")
        st.write(
            "Only your Sandbox API Key and API Secret are required. The app captures profile fields when the "
            "provider returns them, including turnover, gross total income, and e-invoice status."
        )
    with cred2:
        st.markdown("**GSTR-1 / GSTR-3B — needs taxpayer consent**")
        st.write(
            "Requires an OTP consent flow for that specific GSTIN. We never ask for or store the GST portal "
            "password — only the Sandbox API key/secret, the GSTIN, the portal username, and the OTP the "
            "taxpayer receives on their registered mobile or email."
        )
    st.caption("Taxpayer sessions live only in this browser tab's memory and may expire.")

tab_input, tab_run, tab_dashboard, tab_otp, tab_gstr, tab_log = st.tabs(
    [
        "① Input GSTINs",
        "② Run Lookup",
        "③ Dashboard & Export",
        "④ Taxpayer OTP Auth",
        "⑤ GSTR-1 / GSTR-3B Pull",
        "⑥ Activity Log",
    ]
)

# ------------------------------------------------------------
# Tab 1: Input GSTINs
# ------------------------------------------------------------
with tab_input:
    st.subheader("Input multiple GSTINs")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**Paste GSTINs**")
        pasted = st.text_area("Paste GSTINs, one per line", height=200, placeholder="27ABCDE1234F1Z5\n07ABCDE1234F1Z2", label_visibility="collapsed")
        if st.button("➕ Add pasted GSTINs", disabled=not pasted.strip()):
            added = {clean_gstin(line) for line in pasted.splitlines() if clean_gstin(line)}
            if added:
                st.session_state["gstin_list"] = sorted(set(st.session_state.get("gstin_list", [])) | added)
                log_activity("GSTINs added", f"{len(added)} GSTIN(s) added via paste", level="info")
                st.toast(f"Added {len(added)} GSTIN(s).", icon="➕")
                st.rerun()
            else:
                st.warning("No GSTIN-looking text found in the pasted content.")

    with col2:
        st.markdown("**Upload CSV / Excel**")
        uploaded = st.file_uploader("Upload file", type=["csv", "xlsx", "xls"], label_visibility="collapsed")
        if uploaded is not None:
            try:
                df_upload = read_uploaded_gstins(uploaded)
                with st.expander("Preview uploaded file", expanded=False):
                    st.dataframe(df_upload.head(10), use_container_width=True)
                columns = list(df_upload.columns)
                default_idx = next((i for i, c in enumerate(columns) if "gst" in str(c).lower()), 0)
                selected_col = st.selectbox("GSTIN column", columns, index=default_idx)
                if st.button("➕ Add from file"):
                    added = {clean_gstin(v) for v in df_upload[selected_col].dropna().astype(str).tolist() if clean_gstin(v)}
                    if added:
                        st.session_state["gstin_list"] = sorted(set(st.session_state.get("gstin_list", [])) | added)
                        log_activity("GSTINs added", f"{len(added)} GSTIN(s) added from file", level="info")
                        st.toast(f"Added {len(added)} GSTIN(s) from file.", icon="➕")
                        st.rerun()
                    else:
                        st.warning("No GSTIN-looking values found in that column.")
            except Exception as exc:
                st.error(f"Could not read uploaded file: {exc}")

    st.divider()
    gstin_list = st.session_state.get("gstin_list", [])
    st.markdown(f"**Working list** — {len(gstin_list)} GSTIN(s)")

    mc1, mc2, mc3 = st.columns([2, 1, 1])
    with mc1:
        manual_add = st.text_input("Add a single GSTIN", placeholder="27ABCDE1234F1Z5", label_visibility="collapsed")
    with mc2:
        if st.button("➕ Add", use_container_width=True, disabled=not manual_add.strip()):
            cleaned = clean_gstin(manual_add)
            if cleaned:
                st.session_state["gstin_list"] = sorted(set(gstin_list) | {cleaned})
                log_activity("GSTIN added", cleaned, level="info")
                st.rerun()
    with mc3:
        if st.button("🗑️ Clear list", use_container_width=True, disabled=not gstin_list):
            st.session_state["gstin_list"] = []
            log_activity("GSTIN list cleared", f"{len(gstin_list)} GSTIN(s) removed", level="warning")
            st.rerun()

    gstin_list = st.session_state.get("gstin_list", [])

    if not gstin_list:
        st.info("Paste GSTINs, upload a file, or add one manually to get started.")
    else:
        valid_count = sum(1 for g in gstin_list if gstin_format_valid(g))
        checksum_count = sum(1 for g in gstin_list if gstin_format_valid(g) and gstin_checksum_valid(g))

        m1, m2, m3 = st.columns(3)
        m1.metric("Total GSTINs", len(gstin_list))
        m2.metric("Valid format", valid_count)
        m3.metric("Checksum valid", checksum_count)

        preview_df = pd.DataFrame(
            [
                {
                    "GSTIN": g,
                    "Valid Format": "✅" if gstin_format_valid(g) else "❌",
                    "Checksum Valid": "✅" if (gstin_format_valid(g) and gstin_checksum_valid(g)) else "❌",
                    "State": gstin_state(g),
                    "PAN": pan_from_gstin(g),
                }
                for g in gstin_list
            ]
        )

        search_q = st.text_input("🔍 Filter by GSTIN, state, or PAN", key="input_filter")
        st.dataframe(filter_dataframe(preview_df, search_q), use_container_width=True, hide_index=True)

        invalid = [g for g in gstin_list if not gstin_format_valid(g)]
        bad_checksum = [g for g in gstin_list if gstin_format_valid(g) and not gstin_checksum_valid(g)]
        if invalid or bad_checksum:
            with st.expander(f"⚠️ {len(invalid) + len(bad_checksum)} GSTIN(s) need attention before running", expanded=False):
                if invalid:
                    st.write(f"Invalid format ({len(invalid)}): " + ", ".join(invalid))
                if bad_checksum:
                    st.write(f"Checksum mismatch ({len(bad_checksum)}): " + ", ".join(bad_checksum))
                st.caption("These will be marked Skipped/Warning automatically during the bulk run — fix them above if they were mistyped.")

        to_remove = st.multiselect("Remove specific GSTINs from the list", options=gstin_list, key="remove_select")
        if to_remove and st.button("Remove selected"):
            st.session_state["gstin_list"] = sorted(set(gstin_list) - set(to_remove))
            log_activity("GSTINs removed", f"{len(to_remove)} GSTIN(s) removed", level="info")
            st.rerun()

# ------------------------------------------------------------
# Tab 2: Run Lookup
# ------------------------------------------------------------
with tab_run:
    st.subheader("Run bulk public API lookup")

    gstin_list = st.session_state.get("gstin_list", [])
    customers_now = st.session_state.get("customers_df", pd.DataFrame(columns=CUSTOMER_COLUMNS))

    if not gstin_list:
        st.info("Add GSTINs in the **① Input GSTINs** tab first.")
    elif not api_key or not api_secret:
        st.warning("Enter your Sandbox API Key and API Secret in the sidebar.")
    else:
        to_process, cached = partition_targets(gstin_list, customers_now, force_refresh)

        c1, c2, c3 = st.columns(3)
        c1.metric("Total GSTINs", len(gstin_list))
        c2.metric("Already fetched", len(cached))
        c3.metric("Will call API now", len(to_process))

        if cached and not force_refresh:
            st.caption(f"{len(cached)} GSTIN(s) already have a successful result and will be skipped. Turn on **Force refresh** in the sidebar to re-fetch them.")

        if st.button("🚀 Start bulk lookup", type="primary", disabled=not to_process):
            execute_bulk_run(
                targets=to_process,
                api_key=api_key,
                api_secret=api_secret,
                base_url=base_url,
                financial_year=financial_year,
                fetch_profile=fetch_profile,
                fetch_filings=fetch_filings,
                fetch_preference=fetch_preference,
                accept_cache=accept_cache,
                max_workers=max_workers,
                delay_seconds=delay_seconds,
            )

        if not to_process and gstin_list:
            st.success("Everything in your list already has a successful result. Nothing to do.")

# ------------------------------------------------------------
# Tab 3: Dashboard & Export
# ------------------------------------------------------------
with tab_dashboard:
    st.subheader("Dashboard & export")

    customers = st.session_state.get("customers_df", pd.DataFrame(columns=CUSTOMER_COLUMNS))
    filings = st.session_state.get("filings_df", pd.DataFrame(columns=FILING_COLUMNS))
    errors = st.session_state.get("errors_df", pd.DataFrame(columns=ERROR_COLUMNS))
    return_responses = st.session_state.get("return_responses_df", pd.DataFrame(columns=RETURN_RESPONSE_COLUMNS))
    return_facts = st.session_state.get("return_facts_df", pd.DataFrame(columns=RETURN_FACT_COLUMNS))
    gstr1_invoices = st.session_state.get("gstr1_invoices_df", pd.DataFrame(columns=GSTR1_INVOICE_COLUMNS))

    if customers.empty:
        st.info("No results yet — run the bulk lookup in the **② Run Lookup** tab first.")
    else:
        failed_gstins = customers.loc[customers[COL_API_STATUS].isin(FAILURE_API_STATUSES), COL_GSTIN].tolist()

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("GSTINs processed", len(customers))
        m2.metric("Active GSTINs", int(customers[COL_GSTIN_STATUS].astype(str).str.contains("Active", case=False, na=False).sum()))
        m3.metric("Turnover found", int(customers["Aggregate Turnover"].astype(str).str.strip().ne("").sum()) if "Aggregate Turnover" in customers.columns else 0)
        m4.metric("Public filing rows", len(filings))
        m5.metric("GSTR responses", len(return_responses))
        m6.metric("Failed / needs retry", len(failed_gstins))

        if failed_gstins and st.button(f"🔁 Retry {len(failed_gstins)} failed GSTIN(s)"):
            execute_bulk_run(
                targets=failed_gstins,
                api_key=api_key,
                api_secret=api_secret,
                base_url=base_url,
                financial_year=financial_year,
                fetch_profile=fetch_profile,
                fetch_filings=fetch_filings,
                fetch_preference=fetch_preference,
                accept_cache=accept_cache,
                max_workers=max_workers,
                delay_seconds=delay_seconds,
            )

        st.markdown("#### 📈 Analytics overview")
        cc1, cc2 = st.columns(2)
        with cc1:
            render_status_pie(customers, COL_GSTIN_STATUS, "GSTIN status mix")
        with cc2:
            render_state_bar(customers, "State From GSTIN", "GSTINs by state")

        cc3, cc4 = st.columns(2)
        with cc3:
            render_status_pie(customers, "Filing Frequency", "Filing frequency mix")
        with cc4:
            render_status_pie(customers, COL_API_STATUS, "API call outcome")

        st.markdown("#### Taxpayer details")
        search_dashboard = st.text_input("🔍 Filter taxpayer details", key="dashboard_filter", placeholder="Search by GSTIN, name, state, status…")
        display_customers = decorate_for_display(customers, [COL_API_STATUS, COL_GSTIN_STATUS])
        st.dataframe(filter_dataframe(display_customers, search_dashboard), use_container_width=True, hide_index=True)

        st.markdown("#### Turnover, GTI & e-invoice summary")
        turnover_cols = [
            "GSTIN", "Legal Name of Business", "Trade Name", "Aggregate Turnover", "Aggregate Turnover FY",
            "Gross Total Income", "Gross Total Income FY", "E-Invoice Status", "Turnover Data Status",
            "Filing Frequency", COL_GSTIN_STATUS,
        ]
        existing_turnover_cols = [c for c in turnover_cols if c in customers.columns]
        st.dataframe(customers[existing_turnover_cols], use_container_width=True, hide_index=True)

        with st.expander("Debug — why turnover may be blank"):
            st.write(
                "If Aggregate Turnover or Gross Total Income is blank, the provider's API response simply did not "
                "include those fields for that GSTIN. The columns below show any related keys that were detected, "
                "plus the raw profile JSON, so you can confirm the exact field names returned."
            )
            debug_cols = ["GSTIN", "Turnover Data Status", "Detected Turnover/E-Invoice Keys", "Raw GSTIN Profile JSON"]
            existing_debug_cols = [c for c in debug_cols if c in customers.columns]
            st.dataframe(customers[existing_debug_cols], use_container_width=True, hide_index=True)

        st.markdown("#### Public filing table")
        search_filings = st.text_input("🔍 Filter filing table", key="filings_filter", placeholder="Search by GSTIN, ARN, status…")
        st.dataframe(filter_dataframe(filings, search_filings), use_container_width=True, hide_index=True)

        if not gstr1_invoices.empty:
            st.markdown("#### GSTR-1 invoice rows")
            st.dataframe(gstr1_invoices, use_container_width=True, hide_index=True)

        if not return_facts.empty:
            st.markdown("#### Return fact rows")
            st.dataframe(return_facts, use_container_width=True, hide_index=True)

        if not return_responses.empty:
            st.markdown("#### Raw return API responses")
            st.dataframe(return_responses.drop(columns=["Raw JSON"], errors="ignore"), use_container_width=True, hide_index=True)

        if not errors.empty:
            st.markdown("#### Errors / failed GSTINs")
            st.dataframe(errors, use_container_width=True, hide_index=True)

        st.markdown("#### Export")
        excel_bytes = make_excel(customers, filings, errors, return_responses, return_facts, gstr1_invoices)
        json_bytes = make_json_export(customers, filings, errors, return_facts, gstr1_invoices)
        stamp = datetime.now().strftime("%Y%m%d_%H%M")

        ex1, ex2, ex3, ex4 = st.columns(4)
        with ex1:
            st.download_button(
                "📊 Full Excel report",
                data=excel_bytes,
                file_name=f"gst_compliance_report_{stamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        with ex2:
            st.download_button(
                "Taxpayer CSV",
                data=customers.to_csv(index=False).encode("utf-8"),
                file_name=f"taxpayer_details_{stamp}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with ex3:
            st.download_button(
                "Filing table CSV",
                data=filings.to_csv(index=False).encode("utf-8"),
                file_name=f"filing_table_{stamp}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with ex4:
            st.download_button(
                "Full JSON export",
                data=json_bytes,
                file_name=f"gst_compliance_export_{stamp}.json",
                mime="application/json",
                use_container_width=True,
            )

# ------------------------------------------------------------
# Tab 4: Taxpayer OTP Auth
# ------------------------------------------------------------
with tab_otp:
    st.subheader("Taxpayer OTP authentication")
    st.warning("Required only for private GSTR-1 / GSTR-3B data. Never asks for the GST portal password.")

    customers = st.session_state.get("customers_df", pd.DataFrame(columns=CUSTOMER_COLUMNS))
    gstin_options = sorted(
        set(st.session_state.get("gstin_list", []))
        | set(customers["GSTIN"].dropna().tolist() if not customers.empty else [])
    )

    step1, step2 = st.columns(2)

    with step1:
        st.markdown("##### Step 1 — Generate OTP")
        if not gstin_options:
            manual_gstin = st.text_input("GSTIN", placeholder="Enter GSTIN")
            auth_gstin = clean_gstin(manual_gstin)
        else:
            auth_gstin = st.selectbox("Select GSTIN", gstin_options)

        if auth_gstin:
            st.markdown(render_gstin_anatomy(auth_gstin), unsafe_allow_html=True)

        username = st.text_input("GST portal username for this GSTIN", placeholder="Portal username, not password")
        st.caption("The customer should enable API access on the GST portal before requesting an OTP.")

        if st.button("Generate OTP", type="primary"):
            if not api_key or not api_secret:
                st.error("Enter your Sandbox API Key and API Secret in the sidebar first.")
            elif not auth_gstin or not username:
                st.error("Enter a GSTIN and the GST portal username.")
            else:
                try:
                    app_token = authenticate(api_key, api_secret, base_url)
                    payload = generate_taxpayer_otp(base_url, api_key, app_token, auth_gstin, username)
                    st.session_state["pending_taxpayer_auth"] = {
                        "gstin": auth_gstin,
                        "username": username,
                        "app_token": app_token,
                        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    }
                    log_activity("OTP generated", f"GSTIN {auth_gstin}", level="success")
                    st.success("OTP sent. Ask the customer for the code from their registered mobile/email.")
                    with st.expander("Raw API response"):
                        st.json(payload)
                except Exception as exc:
                    log_activity("OTP generation failed", str(exc), level="error")
                    st.error(f"OTP generation failed: {exc}")

    with step2:
        st.markdown("##### Step 2 — Verify OTP")
        pending = st.session_state.get("pending_taxpayer_auth", {})
        if not pending:
            st.caption("Generate an OTP on the left to continue.")
        else:
            st.markdown(f'Pending GSTIN · <span class="chip mono">{html.escape(str(pending.get("gstin", "")))}</span>', unsafe_allow_html=True)
            st.caption(f"Username: {pending.get('username', '')} · Requested at {pending.get('generated_at', '')}")
            otp = st.text_input("Enter OTP", type="password")

            if st.button("Verify OTP and start session", type="primary"):
                try:
                    payload = verify_taxpayer_otp(base_url, api_key, pending["app_token"], pending["gstin"], pending["username"], otp)
                    taxpayer_token = get_nested(payload, "data", "access_token")
                    session_expiry = get_nested(payload, "data", "session_expiry", default="")
                    if not taxpayer_token:
                        st.error("The OTP-verified response did not contain a taxpayer access token.")
                        with st.expander("Raw API response"):
                            st.json(payload)
                    else:
                        sessions = st.session_state.get("taxpayer_sessions", {})
                        sessions[pending["gstin"]] = {
                            "gstin": pending["gstin"],
                            "username": pending["username"],
                            "taxpayer_token": taxpayer_token,
                            "session_expiry": session_expiry,
                            "verified_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        }
                        st.session_state["taxpayer_sessions"] = sessions
                        st.session_state["pending_taxpayer_auth"] = {}
                        log_activity("Taxpayer session started", f"GSTIN {pending['gstin']}", level="success")
                        st.success("Taxpayer session started. You can now pull GSTR-1 / GSTR-3B data.")
                        st.toast("Taxpayer session active.", icon="🔐")
                except Exception as exc:
                    log_activity("OTP verification failed", str(exc), level="error")
                    st.error(f"OTP verification failed: {exc}")

    sessions = st.session_state.get("taxpayer_sessions", {})
    if sessions:
        st.divider()
        st.markdown("##### Active taxpayer sessions")
        session_preview = pd.DataFrame(
            [
                {"GSTIN": k, "Username": v.get("username", ""), "Verified At": v.get("verified_at", ""), "Session Expiry": v.get("session_expiry", "")}
                for k, v in sessions.items()
            ]
        )
        st.dataframe(session_preview, use_container_width=True, hide_index=True)
        if st.button("Clear all taxpayer sessions"):
            st.session_state["taxpayer_sessions"] = {}
            log_activity("Sessions cleared", f"{len(sessions)} session(s) cleared", level="warning")
            st.rerun()

# ------------------------------------------------------------
# Tab 5: Pull GSTR-1 / GSTR-3B
# ------------------------------------------------------------
with tab_gstr:
    st.subheader("Pull GSTR-1 / GSTR-3B data")

    sessions = st.session_state.get("taxpayer_sessions", {})
    filings = st.session_state.get("filings_df", pd.DataFrame(columns=FILING_COLUMNS))

    if not sessions:
        st.info("Complete **④ Taxpayer OTP Auth** for at least one GSTIN first.")
    else:
        session_gstins = sorted(sessions.keys())
        selected_gstin = st.selectbox("Authenticated GSTIN", session_gstins)
        st.markdown(render_gstin_anatomy(selected_gstin), unsafe_allow_html=True)

        gstin_filings = filings[filings["GSTIN"] == selected_gstin] if not filings.empty else pd.DataFrame(columns=FILING_COLUMNS)

        period_options: List[str] = []
        period_map: Dict[str, Tuple[str, str]] = {}
        for _, row in gstin_filings.iterrows():
            ret_period = str(row.get("Return Period", ""))
            year_p, month_p = parse_period_to_year_month(ret_period)
            if year_p and month_p:
                label = f"{row.get('Return Type', '')} | {ret_period} | Filed {row.get('Date of Filing', '')} | ARN {row.get('ARN', '')}"
                period_options.append(label)
                period_map[label] = (year_p, month_p)

        use_from_filing = st.checkbox("Select period from public filing table", value=bool(period_options))

        if use_from_filing and period_options:
            selected_period = st.selectbox("Filed return period", period_options)
            year, month = period_map[selected_period]
            st.caption(f"Selected period: year {year}, month {month}")
        else:
            pc1, pc2 = st.columns(2)
            with pc1:
                year = st.text_input("Year", value=str(datetime.now().year))
            with pc2:
                month = st.text_input("Month", value=f"{datetime.now().month:02d}", help="01 to 12")

        st.markdown("**What to fetch**")
        wc1, wc2 = st.columns(2)
        with wc1:
            fetch_gstr3b = st.checkbox("GSTR-3B details", value=True)
        with wc2:
            fetch_gstr1 = st.checkbox("GSTR-1 documents", value=True)

        selected_sections: List[str] = []
        if fetch_gstr1:
            sc1, sc2 = st.columns([3, 1])
            with sc2:
                st.write("")
                if st.button("Select all"):
                    st.session_state["gstr1_section_pick"] = list(GSTR1_SECTIONS.keys())
                    st.rerun()
                if st.button("Common only"):
                    st.session_state["gstr1_section_pick"] = list(COMMON_GSTR1_SECTIONS)
                    st.rerun()
            with sc1:
                selected_sections = st.multiselect(
                    "GSTR-1 sections",
                    options=list(GSTR1_SECTIONS.keys()),
                    key="gstr1_section_pick",
                )

        if st.button("📥 Pull return data", type="primary"):
            if not re.match(r"^\d{4}$", str(year)):
                st.error("Year must be YYYY, e.g. 2026.")
            elif not re.match(r"^(0[1-9]|1[0-2])$", str(month)):
                st.error("Month must be 01 to 12.")
            else:
                taxpayer_token = sessions[selected_gstin]["taxpayer_token"]
                response_rows: List[Dict[str, Any]] = []
                fact_frames: List[pd.DataFrame] = []
                invoice_rows: List[Dict[str, Any]] = []
                error_rows: List[Dict[str, Any]] = []

                total_steps = (1 if fetch_gstr3b else 0) + (len(selected_sections) if fetch_gstr1 else 0)
                total_steps = max(total_steps, 1)

                with st.status(f"Pulling return data for {selected_gstin}…", expanded=True) as status_box:
                    progress_bar = st.progress(0.0)
                    done = 0

                    if fetch_gstr3b:
                        try:
                            payload = fetch_gstr3b_details(base_url, api_key, taxpayer_token, year, month)
                            response_rows.append(add_return_response(selected_gstin, "GSTR-3B", "DETAILS", year, month, payload))
                            fact_frames.append(make_return_facts(selected_gstin, "GSTR-3B", "DETAILS", year, month, payload))
                            status_box.write("✅ GSTR-3B details fetched.")
                        except Exception as exc:
                            error_rows.append(build_error(selected_gstin, "GSTR-3B Details", str(exc)))
                            status_box.write(f"❌ GSTR-3B fetch failed: {exc}")
                        done += 1
                        progress_bar.progress(done / total_steps)

                    if fetch_gstr1:
                        for section in selected_sections:
                            slug = GSTR1_SECTIONS[section]
                            try:
                                payload = fetch_gstr1_section(base_url, api_key, taxpayer_token, slug, year, month)
                                response_rows.append(add_return_response(selected_gstin, "GSTR-1", section, year, month, payload))
                                fact_frames.append(make_return_facts(selected_gstin, "GSTR-1", section, year, month, payload))
                                invoice_rows.extend(extract_gstr1_invoice_rows(payload, selected_gstin, year, month, section))
                                status_box.write(f"✅ GSTR-1 {section} fetched.")
                            except Exception as exc:
                                error_rows.append(build_error(selected_gstin, f"GSTR-1 {section}", str(exc)))
                                status_box.write(f"❌ GSTR-1 {section} failed: {exc}")
                            done += 1
                            progress_bar.progress(done / total_steps)

                    status_box.update(label=f"Return data pull completed for {selected_gstin}", state="complete")

                new_responses = pd.DataFrame(response_rows, columns=RETURN_RESPONSE_COLUMNS)
                new_facts = pd.concat(fact_frames, ignore_index=True) if fact_frames else pd.DataFrame(columns=RETURN_FACT_COLUMNS)
                new_invoices = pd.DataFrame(invoice_rows, columns=GSTR1_INVOICE_COLUMNS)
                new_errors = pd.DataFrame(error_rows, columns=ERROR_COLUMNS)

                st.session_state["return_responses_df"] = append_df(st.session_state.get("return_responses_df", pd.DataFrame(columns=RETURN_RESPONSE_COLUMNS)), new_responses, RETURN_RESPONSE_COLUMNS)
                st.session_state["return_facts_df"] = append_df(st.session_state.get("return_facts_df", pd.DataFrame(columns=RETURN_FACT_COLUMNS)), new_facts, RETURN_FACT_COLUMNS)
                st.session_state["gstr1_invoices_df"] = append_df(st.session_state.get("gstr1_invoices_df", pd.DataFrame(columns=GSTR1_INVOICE_COLUMNS)), new_invoices, GSTR1_INVOICE_COLUMNS)
                st.session_state["errors_df"] = append_df(st.session_state.get("errors_df", pd.DataFrame(columns=ERROR_COLUMNS)), new_errors, ERROR_COLUMNS)

                log_activity(
                    "GSTR pull",
                    f"{selected_gstin} · {year}-{month} · {len(response_rows)} response(s), {len(error_rows)} error(s)",
                    level="warning" if error_rows else "success",
                )
                st.toast("Return data pull completed.", icon="📥")

        st.markdown("#### Pulled return data preview")
        return_responses = st.session_state.get("return_responses_df", pd.DataFrame(columns=RETURN_RESPONSE_COLUMNS))
        return_facts = st.session_state.get("return_facts_df", pd.DataFrame(columns=RETURN_FACT_COLUMNS))
        gstr1_invoices = st.session_state.get("gstr1_invoices_df", pd.DataFrame(columns=GSTR1_INVOICE_COLUMNS))

        if not return_responses.empty:
            st.markdown("##### API response summary")
            st.dataframe(return_responses.drop(columns=["Raw JSON"], errors="ignore"), use_container_width=True, hide_index=True)

        if not gstr1_invoices.empty:
            st.markdown("##### GSTR-1 invoice rows")
            st.dataframe(gstr1_invoices, use_container_width=True, hide_index=True)

        if not return_facts.empty:
            st.markdown("##### Generic return facts")
            st.dataframe(return_facts, use_container_width=True, hide_index=True)

        if not return_responses.empty:
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            export_excel = make_excel(
                st.session_state.get("customers_df", pd.DataFrame(columns=CUSTOMER_COLUMNS)),
                st.session_state.get("filings_df", pd.DataFrame(columns=FILING_COLUMNS)),
                st.session_state.get("errors_df", pd.DataFrame(columns=ERROR_COLUMNS)),
                return_responses,
                return_facts,
                gstr1_invoices,
            )
            st.download_button(
                "📊 Download return data Excel",
                data=export_excel,
                file_name=f"gstr1_gstr3b_data_{selected_gstin}_{stamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

# ------------------------------------------------------------
# Tab 6: Activity Log
# ------------------------------------------------------------
with tab_log:
    st.subheader("Activity log")
    log = st.session_state.get("activity_log", [])

    if not log:
        st.info("Nothing logged yet. Connection tests, bulk runs, OTP steps, and GSTR pulls will appear here.")
    else:
        level_filter = st.multiselect(
            "Filter by level",
            options=["success", "info", "warning", "error"],
            default=["success", "info", "warning", "error"],
        )
        filtered_log = [entry for entry in log if entry["Level"] in level_filter]

        level_classes = {"success": "badge-ok", "info": "badge-idle", "warning": "badge-warn", "error": "badge-bad"}
        level_icons = {"success": "✅", "info": "ℹ️", "warning": "⚠️", "error": "❌"}

        rows_html = []
        for entry in filtered_log:
            cls = level_classes.get(entry["Level"], "badge-idle")
            icon = level_icons.get(entry["Level"], "•")
            action = html.escape(str(entry.get("Action", "")))
            detail = html.escape(str(entry.get("Detail", "")))
            time_str = html.escape(str(entry.get("Time", "")))
            detail_html = f'<div class="log-detail">{detail}</div>' if detail else ""
            row = (
                '<div class="log-row">'
                f'<span class="badge {cls}">{icon} {entry["Level"]}</span>'
                f'<div class="log-body"><b>{action}</b>{detail_html}</div>'
                f'<span class="log-time mono">{time_str}</span>'
                "</div>"
            )
            rows_html.append(row)
        st.markdown('<div class="log-list">' + "".join(rows_html) + "</div>", unsafe_allow_html=True)

        log_df = pd.DataFrame(filtered_log)
        st.download_button(
            "Download activity log (CSV)",
            data=log_df.to_csv(index=False).encode("utf-8"),
            file_name=f"activity_log_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv",
        )

st.markdown(
    """
    <div class="app-footer">
        GSTIN Compliance Suite · Powered by the Sandbox GST Compliance API · Built with Streamlit<br>
        This tool never requests or stores GST portal passwords. Private return data requires the taxpayer's own OTP consent for each GSTIN.
    </div>
    """,
    unsafe_allow_html=True,
)
